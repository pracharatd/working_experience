from openpyxl import load_workbook
import pyodbc as odbc

#EXCEL
workbook = load_workbook('imported_01.xlsx')
sheet = workbook.active
values = []


for row in sheet.iter_rows(min_row = 2, values_only = True):
    values.append(row)

#DATABASE

DRIVER_NAME = 'SQL Server'
SERVER_NAME = 'NGDNB620013\SQLEXPRESS'
DATABASE_NAME = 'testsql'

conn_string = f"""
    Driver={{{DRIVER_NAME}}};
    Server={SERVER_NAME};
    Database={DATABASE_NAME};
    Trust_Connection=yes;
    username=practicesql;
    password=Best0681;
"""
conn = odbc.connect(conn_string)
print(conn)

sql = '''
        INSERT INTO Table_1 (title, price, is_necessary)
        VALUES (?, ?, ?)
'''

cursor = conn.cursor()
cursor.executemany(sql,values)
conn.commit()