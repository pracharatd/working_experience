from openpyxl import load_workbook

import pyodbc as odbc

#EXCEL
workbook = load_workbook('imported_01.xlsx')
sheet = workbook.active
values = []

for row in sheet.iter_rows(min_row = 2, values_only = True):
    values.append(row)

#DATABASE
conn_string = f"""
    Driver={{{'SQL Server'}}};
    Server={'wuttbestpracticesql.database.windows.net'};
    Database={'bestsqlwutt'};
    Trust_Connection=yes;
    uid=pracharat.d;
    pwd=Best0681;
"""
conn = odbc.connect(conn_string)

sql = '''
        INSERT INTO Table_1 (title, price, is_necessary)
        VALUES (?, ?, ?)
'''

cursor = conn.cursor()
cursor.executemany(sql,values)
conn.commit()